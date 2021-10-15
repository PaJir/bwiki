# 装备数据
import sqlite3
import io
import openpyxl

db_name = "./redive_cn0925.db"
db_name_jp = "./redive_jp0925.db"

conn = sqlite3.connect(db_name)
cursor = conn.cursor()
conn_jp = sqlite3.connect(db_name_jp)
cursor_jp = conn_jp.cursor()
new_craft = ["太阳剑", "公主短剑", "狮子王的护臂", "天使弓", "女王之矛", "生命法杖", "太阳法杖", "元素之心", 
    "新月的悲叹", "替身手镯", "冰之大剑", "鹰神之煌剑", "暗之刃", "阿尔忒弥斯之弓", "战女王的烈斧", "司法神的锡杖", 
    "太阳护身符", "龙之泪", "月之手镯", "混沌之刃", "苍雷拳", "精灵树之弓", "盖亚之矛", "水精灵之斧", "大天使法杖", 
    "焰帝戒指", "大贤者宝珠", "圣信徒吊坠", "猫神吊坠", "闪电之刃", "百华白樱刀", "翡翠匕首", "冰霜矛", "火精灵法杖", 
    "异界法杖", "千禧耳环", "海神耳饰", "精灵王护石", "流祸苍刃", "深红爪", "深渊之弓", "神判圣斧", "创世法杖", "赫怒龙吊坠", 
    "圣兽的祈祷", "人鱼公主的灵泪", "提尔纳诺短剑", "红天弓", "护天的圣枪", "斗神勇腕", "睿智手镯", "英盾王的手镯", 
    "狮子王的守护", "珍贵匕首", "火焰护臂", "雷光弓", "绯龙矛", "月华法杖", "天使护甲", "女王盾", "圣骑士护胫甲", 
    "巫师头巾", "羽翼头盔", "光辉剑", "地狱火之斧", "秘银铠甲", "绯红铠甲", "圣人长袍", "隐士服", "天使鞋", "法王头巾", 
    "暴雪之爪", "兴风者", "凤凰之杖", "十字军盔甲", "紫罗兰盔甲", "苍天之袍", "守护者之盾", "死灵法师之靴", "福音之冠", 
    "凤凰刀", "独角兽匕首", "大恶魔法杖", "猩红铠甲", "弑龙铠甲", "翠绿灵衣", "暗黑礼裙", "妖精鞋", "宝瓶鞋", "月形拳套", 
    "白银大翼弓", "死灵斧", "幻影铠甲", "苍辉之铠", "纯洁巫女服", "神盾战衣", "星咏圆盾", "忘哭之冠", "天开圣剑", "熔岩短剑", 
    "风神矛", "神花圣杖", "九天之铠", "圣樱之铠", "女皇长袍", "少女服", "地狱胫甲", "白祈圣冠", "罪过的大剑", "妖刀血鸦", 
    "悲叹河之爪", "盖亚之斧", "亚特兰蒂斯之杖", "雷霆之杖", "愤怒女王的礼裙", "煌金王铠", "魔导王长袍", "鬼焰斗衣", 
    "血红宝石高跟鞋", "皇家守卫帽"]

def write_same(data):
    s = ""
    s += str(data[1]) + " "
    for i in range(37):
        s += str(data[i]).replace("\\n","<br/>").replace(" ", "") + " "
    s += str(lowest_rank(data[0])) + " "
    s += str(lowest_map(data[0])) + " "
    return s
# 初始RANK
def lowest_rank(id):
    id = str(id)
    sql_rank = """
select promotion_level pl
from unit_promotion
where equip_slot_1 = '""" + id + "'or equip_slot_2 = '" + id + "' or equip_slot_3 = '" + id + "' or equip_slot_4 = '" + id + "' or equip_slot_5 = '" + id + "' or equip_slot_6 = '" + id + """'
order by pl ASC
LIMIT 1"""
    cursor_jp.execute(sql_rank)
    all_data = cursor_jp.fetchone()
    assert all_data != None
    return all_data[0]
#初始地图
def lowest_map(id):
    id = str(id)
    wb = openpyxl.load_workbook("./read_db_equipment.xlsx")
    sheetname = wb.sheetnames[0]
    sheet = wb[sheetname]
    rows = sheet.max_row
    for row in range(2, rows + 1):
        m = sheet.cell(row, 1).value
        c = sheet.cell(row, 3).value
        if id[2:] in c:
            return m.split("-")[0]
    return "badbadbadbadbadofmap"
def equipment_craft():
    f = open("equip_data_craft_rate.txt", "w")
    # 7 + 15 + 15 + 21
    sql_main = """
select d.equipment_id, d.equipment_name, d.description ddes, r.description rdes, d.promotion_level, d.sale_price, d.require_level, 
d.hp, d.atk, d.magic_str, d.def, d.magic_def, d.physical_critical, d.magic_critical, d.wave_hp_recovery, d.wave_energy_recovery, d.dodge, d.life_steal, d.hp_recovery_rate, d.energy_recovery_rate, d.energy_reduce_rate, d.accuracy, 
r.hp, r.atk, r.magic_str, r.def, r.magic_def, r.physical_critical, r.magic_critical, r.wave_hp_recovery, r.wave_energy_recovery, r.dodge, r.life_steal, r.hp_recovery_rate, r.energy_recovery_rate, r.energy_reduce_rate, r.accuracy, 
c.crafted_cost, c.condition_equipment_id_1, c.consume_num_1, c.condition_equipment_id_2, c.consume_num_2, c.condition_equipment_id_3, c.consume_num_3, c.condition_equipment_id_4, c.consume_num_4, c.condition_equipment_id_5, c.consume_num_5, c.condition_equipment_id_6, c.consume_num_6, c.condition_equipment_id_7, c.consume_num_7, c.condition_equipment_id_8, c.consume_num_8, c.condition_equipment_id_9, c.consume_num_9, c.condition_equipment_id_10, c.consume_num_10 
from equipment_data d join equipment_enhance_rate r join equipment_craft c on d.equipment_id = c.equipment_id and c.equipment_id = r.equipment_id"""
    cursor.execute(sql_main)
    all_data = cursor.fetchall()
    # 下面这个为没有craft的装备定制
    sql_main2 = """
select d.equipment_id, d.equipment_name, d.description ddes, r.description rdes, d.promotion_level, d.sale_price, d.require_level, 
d.hp, d.atk, d.magic_str, d.def, d.magic_def, d.physical_critical, d.magic_critical, d.wave_hp_recovery, d.wave_energy_recovery, d.dodge, d.life_steal, d.hp_recovery_rate, d.energy_recovery_rate, d.energy_reduce_rate, d.accuracy, 
r.hp, r.atk, r.magic_str, r.def, r.magic_def, r.physical_critical, r.magic_critical, r.wave_hp_recovery, r.wave_energy_recovery, r.dodge, r.life_steal, r.hp_recovery_rate, r.energy_recovery_rate, r.energy_reduce_rate, r.accuracy 
from equipment_data d join equipment_enhance_rate r on d.equipment_id = r.equipment_id"""
    cursor.execute(sql_main2)
    low_data = cursor.fetchall()
    for data in all_data:
        f.write(write_same(data))
        # 合成价格
        f.write(str(data[37]) + " ")
        # 合成材料
        craft = ""
        # prob = ""
        for i in range(38, 58, 2):
            if str(data[i]) == "0":
                break
            craft += str(data[i]) + ":" + str(data[i+1]) + ","
            # prob += str(data[i+1]) + ","
        f.write(craft[:-1]  + " \n")

    for data in low_data:
        # if data[1] not in new_craft:
        #     continue
        is_low = True
        for d in all_data:
            if data[0] == d[0]:
                is_low = False
                break
        if not is_low:
            continue
        f.write(write_same(data))
        f.write("\n")
    f.close()

equipment_craft()
conn.close()
conn_jp.close()