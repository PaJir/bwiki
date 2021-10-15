# 输出模块，将数据报表保存为xlsx文件
# 安装依赖包 pip install openpyxl（如果报错的话）
import openpyxl
from openpyxl.cell import MergedCell
colors = ['',
          'style="background-color: lightgreen" ',
          'style="background-color: yellow" ',
          'style="background-color: yellow" ',
          'style="background-color: yellow" ',
          'style="background-color: yellow" ',
          'style="background-color: pink" ',
          'style="background-color: pink" ',
          'style="background-color: pink" ',
          'style="background-color: lightgreen" ',
          'style="background-color: lightgreen" ',
          'style="background-color: lightyellow" ',
          'style="background-color: orange" ',
          'style="background-color: lightblue" ',
          'style="background-color: lightblue" ',
          'style="background-color: lightblue" ',
          'style="background-color: lightblue" ',
          'style="background-color: lightblue" ',
          'style="background-color: lightblue" ',
          'style="background-color: lightblue" ',
          'style="background-color: lightblue" ',
          'style="background-color: lightgray" ',
          'style="background-color: lightgray" ',
          'style="background-color: lightgray" ',
          'style="background-color: lightgray" ',
          'style="background-color: lightgray" ',
          'style="background-color: lightgray" ',
          'style="background-color: lightgray" ',
          'style="background-color: lightgray" ']

laters = ['',
    '|-style="background-color: yellow"\n!rowspan=4|加倍<br>起止默认<br>5:00~4:59\n',
    '|-\n',
    '|-\n',
    '|-\n',
    '|-style="background-color: pink"\n!rowspan=3|卡池\n',
    '|-\n',
    '|-\n',
    '|-style="background-color: lightgreen"\n!rowspan=2|活动\n',
    '|-\n',
    '|-style="background-color: lightyellow"\n!公会战\n',
    '|-style="background-color: orange"\n!露娜塔\n',
    '|-style="background-color: lightblue"\n!rowspan=8|特殊活动\n',
    '|-\n',
    '|-\n',
    '|-\n',
    '|-\n',
    '|-\n',
    '|-\n',
    '|-\n',
    '|-style="background-color: lightgray"\n!rowspan=8|其他\n',
    '|-\n',
    '|-\n',
    '|-\n',
    '|-\n',
    '|-\n',
    '|-\n',
    '|-\n',
    '|}\n</div>',
    ]

def read_xlsx_output_txt():
    wb = openpyxl.load_workbook("./history2.xlsx")
    sheetname = wb.sheetnames[1]
    sheet = wb[sheetname]
    rows = sheet.max_row
    cols = sheet.max_column
    print(f"行数：{rows}")
    print(f"列数：{cols}")
    with open("./history.txt", "w", encoding="utf-8") as wf:
        for row in range(1, rows + 1):
            # wf.write("第"+str(row)+"行\n")
            col = 1
            while col <= cols:
                c = sheet.cell(row, col)
                if isinstance(c, MergedCell):
                    col += 1
                val = c.value
                col_sum = 1
                for j in range(col + 1, cols + 1):
                    if isinstance(sheet.cell(row, j), MergedCell):
                        col_sum += 1
                    # 让空格子连续
                    elif val is None and sheet.cell(row, j).value is None:
                        col_sum += 1
                    else:
                        break
                sum = (col + col_sum - 1) // 4 - (col - 1) // 4
                # if row == 17:
                    # print(row, col, col_sum, sum, val)
                if sum == 0:
                    pass
                elif val is None:
                    wf.write("|colspan=" + str(sum) + "|\n")
                else:
                    wf.write("|" + colors[row] + "colspan=" + str(sum) + "|" + str(val) + "\n")
                col = col + col_sum
            wf.write(laters[row])


read_xlsx_output_txt()
