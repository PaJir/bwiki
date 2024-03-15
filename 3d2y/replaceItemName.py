import os
import shutil
import numpy as np
from PIL import Image
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import imagehash

images_folder = "images"

def compare_images(img1, img2):
    hash1 = imagehash.average_hash(img1)
    hash2 = imagehash.average_hash(img2)
    diff = hash1 - hash2
    return diff

def main():
    wb = load_workbook("./海贼王.xlsx")
    sheetname = wb.sheetnames[0]
    sheet = wb[sheetname]
    image_loader = SheetImageLoader(sheet)
    
    rows = sheet.max_row
    for row in range(1, rows + 1):
        name_cn = os.path.join(images_folder, "道具-" + sheet.cell(row, 3).value + ".png")
        if os.path.exists(name_cn):
            continue
        image1 = image_loader.get("B" + str(row))
        for filename in os.listdir(images_folder):
            # if filename.endswith('.gif'):
            if True:
                img_path = os.path.join(images_folder, filename)
                image2 = Image.open(img_path)
                diff = compare_images(image1, image2)
                image2.close()
                if diff < 0.04:  # 根据需要设置阈值
                    print(diff)
                    shutil.move(os.path.join(images_folder, filename), name_cn)
                    break

def getXlsxFile(sheetname, images_folder):
    wb = load_workbook("./海贼王.xlsx")
    sheet = wb[sheetname]
    image_loader = SheetImageLoader(sheet)
    
    rows = sheet.max_row
    for row in range(2, rows + 1):
        name_cn = os.path.join(images_folder, "徽章-" + sheet.cell(row, 3).value + ".png")
        if os.path.exists(name_cn):
            continue
        image1 = image_loader.get("B" + str(row))
        image1.save(name_cn)


def replaceImageName(folder):
    img_map = {
        "Item_Dungeon_005": "金币本额外次数货币",
        "Item_Dungeon_006": "默契本额外次数货币",
        "Item_GuildContribute": "GVE正义挑战币",
        "组9拷贝9": "招募指针礼包（CBT4）",
        "image": "防护型铜质觉醒材料",
        "image2": "防护型银质觉醒材料",
        "image3": "激斗型铜质觉醒材料",
        "image4": "激斗型银质觉醒材料",
        "image5": "辅佐型铜质觉醒材料",
        "image6": "辅佐型银质觉醒材料",
        "image7": "应援型铜质觉醒材料",
        "image8": "应援型银质觉醒材料",
        "image9": "追击型铜质觉醒材料",
        "image10": "追击型银质觉醒材料",
        "image11": "自成型铜质觉醒材料",
        "image12": "自成型银质觉醒材料",
        "Item_PartnerAwaken_Job111": "防护觉醒金币",
        "Item_PartnerAwaken_Job121": "防护觉醒彩色硬币",
        "Item_PartnerAwaken_Job211": "激斗觉醒金币",
        "Item_PartnerAwaken_Job221": "激斗觉醒彩色硬币",
        "Item_PartnerAwaken_Job311": "辅佐觉醒金币",
        "Item_PartnerAwaken_Job321": "辅佐觉醒彩色硬币",
        "Item_PartnerAwaken_Job411": "应援觉醒金币",
        "Item_PartnerAwaken_Job421": "应援觉醒彩色硬币",
        "Item_PartnerAwaken_Job511": "追击觉醒金币",
        "Item_PartnerAwaken_Job521": "追击觉醒彩色硬币",
        "Item_PartnerAwaken_Job611": "自成觉醒金币",
        "Item_PartnerAwaken_Job621": "自成觉醒彩色硬币",
        # "ui_img_explore__A1": "丛林荒岛A*3个阶级",
        # "ui_img_explore__A2": "丛林荒岛A*3个阶级",
        # "ui_img_explore__A3": "丛林荒岛A*3个阶级",
        # "ui_img_explore__B1": "岩石荒岛A*3个等级",
        # "ui_img_explore__B2": "岩石荒岛A*3个等级",
        # "ui_img_explore__B3": "岩石荒岛A*3个等级",
        # "ui_img_explore__C1": "丛林荒岛B*3个阶级",
        # "ui_img_explore__C2": "丛林荒岛B*3个阶级",
        # "ui_img_explore__C3": "丛林荒岛B*3个阶级",
        # "ui_img_explore__D1": "岩石荒岛B*3个等级",
        # "ui_img_explore__D2": "岩石荒岛B*3个等级",
        # "ui_img_explore__D3": "岩石荒岛B*3个等级",
        # "ui_img_explore__E1": "沙漠荒岛A*3个阶级",
        # "ui_img_explore__E2": "沙漠荒岛A*3个阶级",
        # "ui_img_explore__E3": "沙漠荒岛A*3个阶级",
        # "ui_img_explore__F1": "沙漠荒岛B*3个阶级",
        # "ui_img_explore__F2": "沙漠荒岛B*3个阶级",
        # "ui_img_explore__F3": "沙漠荒岛B*3个阶级",
        "ui_img_explore__G1": "徽章岛",
        "ui_img_explore__G2": "教学岛",
        "ui_img_explore__G3": "特训岛",
        "ui_img_explore__G4": "天空之岛",
        "ui_img_explore__X1": "香波地岛",
        "ui_Item_Explore_icon1": "木箱",
        "ui_Item_Explore_icon2": "铜箱",
        "ui_Item_Explore_icon3": "银箱",
        "ui_Item_Explore_icon4": "金箱",
        "ui_Item_Explore_icon5": "紫箱",
        "ui_Item_Explore_icon7": "铲子",
        "ui_Item_Explore_icon8": "单孔望远镜",
        "ui_Item_Explore_icon9": "圣像",
        "ui_Item_Explore_icon10": "钥匙",
        "Item_Build_Fruit": "秘林之果",
        "Item_Build_Meet": "异兽之肉",
        "Item_Build_Copper": "粗铜",
        "Item_Build_Silver": "秘银",
        "Item_Build_Crystal_1": "晶矿",
        "Item_Build_Crystal_2": "晶石",
        "Item_Build_DesignPaperA": "初级设计图",
        "Item_Build_DesignPaperB": "中级设计图",
        "Item_Build_DesignPaperC": "高级设计图",
        "Item_Build_DesignPaperD": "顶级设计图",
        "1": "6元对应付费钻石",
        "2": "30元对应付费钻石",
        "3": "68元对应付费钻石",
        "128元": "128元对应付费钻石",
        "328元": "328元对应付费钻石",
        "Item_DreamDiamond_01": "30元对应免费钻石",
        "Item_DreamDiamond_02": "68元对应免费钻石",
        "Item_DreamDiamond_03": "128元对应免费钻石",
        "Item_DreamDiamond_04": "198元对应免费钻石",
        "Item_DreamDiamond_05": "328元对应免费钻石",
        "Item_DreamDiamond_06": "648元对应免费钻石",
        "Icon_Treasure_12001": "涡狐之戒",
        "Icon_Treasure_12002": "协力半环",
        "Icon_Treasure_12003": "黑神珠戒",
        "Icon_Treasure_12005": "赤瞳王冠",
        "Icon_Treasure_23005": "开辟黎明",
        "Icon_Treasure_23006": "百褶笛壶",
        "Icon_Treasure_32003": "岚起高号",
        "Icon_Treasure_32005": "精灵之宙",
        "Item_PartnerEvoSpecial_12": "守护者之章",
        "Item_PartnerEvoSpecial_22": "战士之章",
        "Item_PartnerEvoSpecial_32": "指引者之章",
        "Treasure_FJ_R_02": "固质铁壶",
        "Treasure_FJ_SR_01": "蹈雾之壶",
        "Treasure_FJ_SR_03": "密纹高脚杯",
        "Treasure_FJ_SR_05": "幽邃夜杯",
        "Treasure_FJ_SR_08": "权立宝壶",
        "Treasure_FJ_SSR_01": "浊色小盅",
        "Treasure_FJ_SSR_02": "星海盅",
        "Treasure_FJ_SSR_03": "行宫宝壶",
        "Treasure_FJ_SSR_04": "清白酒杯",
        "Treasure_FJ_SSR_05": "巨宝洋瓶",
        "Treasure_FJ_SSR_07": "眸界樽杯",
        "Treasure_FJ_SSR_08": "森隐之壶",
        "Treasure_FJ_SSR_09": "烟藤圣杯",
        "Treasure_PS_R_01": "带环木球",
        "Treasure_PS_R_02": "四角铁盒",
        "Treasure_PS_R_03": "密纹铁球",
        "Treasure_PS_R_05": "三星锥",
        "Treasure_PS_SSR_02": "星蓝徽记",
        "Treasure_PS_SSR_04": "云起镜牌",
        "Treasure_PS_SSR_05": "律海徽记",
        "Treasure_PS_SSR_07": "赤瞑瞑眼",
        "Treasure_PS_SSR_08": "山隐之眼",
        "Treasure_PS_SSR_10": "守御岩鳞",
        "Treasure_TK_R_01": "旧日木杯",
        "Treasure_TK_R_04": "锋刃兽齿",
        "Treasure_TK_SR_01": "钢壁之镯",
        "Treasure_TK_SR_03": "精巧铁镯",
        "Treasure_TK_SR_05": "远闪虚戒",
        "Treasure_TK_SR_08": "弧星圈",
        "Treasure_TK_SSR_01": "渊洄之环",
        "Treasure_TK_SSR_02": "坠星项链",
        "Treasure_TK_SSR_03": "行宫护腕",
        "Treasure_TK_SSR_04": "珠虞象牙",
        "Treasure_TK_SSR_05": "滴海吊坠",
        "Treasure_TK_SSR_07": "匙之坠",
        "Treasure_TK_SSR_08": "金蛇耳坠",
        "Treasure_TK_SSR_10": "鳞钩耳坠",
        "Treasure_WQ_R_01": "泛黄弹弓",
        "Treasure_WQ_SR_01": "黑铁狼牙",
        "Treasure_WQ_SR_08": "星琉杖",
        "Treasure_WQ_SSR_01": "末图黑刃",
        "Treasure_WQ_SSR_02": "星月长刀",
        "Treasure_WQ_SSR_03": "辉光三叉戟",
        "Treasure_WQ_SSR_04": "米光盾徽",
        "Treasure_WQ_SSR_05": "蓝庭魔杖",
        "Treasure_WQ_SSR_06": "跃动手枪",
        "Treasure_WQ_SSR_07": "星瞳权杖",
        "Treasure_WQ_SSR_10": "黑曜石杖",
        "Icon_Buff_ACE_001": "狂暴",
        "Icon_Buff_aim": "提升",
        "Icon_Buff_BeDamDown_02": "得到奥义",
        "Icon_Buff_DonotDrawCard_001": "禁止升级卡牌等级",
        "Icon_Buff_DrawCard_001": "抽牌",
        "Icon_Buff_Kid_001": "基德专属增益",
        "Icon_Buff_lmmune_01": "怪物免疫控制",
        "Icon_Buff_Mihawk_001": "鹰眼专属增益",
        "Icon_DeBuff_candle_01": "Mr3专属1",
        "Icon_DeBuff_combat_01": "禁止使用卡牌/技能",
        "Icon_DeBuff_combat_02": "禁止使用奥义",
        "Icon_DeBuff_DamDown_01": "失去奥义",
        "Icon_DeBuff_DamDown_02": "虚弱",
        "Icon_DeBuff_DonotUpCard_001": "禁止抽牌",
        "Icon_DeBuff_Light_001": "感电",
        "Icon_DeBuff_Mihawk_001": "鹰眼专属减益",
        "Icon_DeBuff_Mr3_01": "Mr3专属2",
        "Icon_DeBuff_Ohm_01": "欧姆专属减益",
        "Icon_DeBuff_Stone_001": "石化",
        "Icon_SupBuff_Charge": "怪物专属通用",
        "Icon_SupBuff_DefUp": "怪物专属通用",
        "Icon_SupBuff_Strength": "怪物专属通用"
    }
    files = os.listdir(folder)
    for file in files:
        new_file = img_map.get(file[:-4], None)
        if new_file is None:
            continue
        print(file, new_file)
        shutil.move(os.path.join(folder, file), os.path.join(folder, new_file+".png"))

def replaceImage(sheetname, old_folder, new_folder, backup_folder):
    wb = load_workbook("./海贼王.xlsx")
    sheet = wb[sheetname]
    # image_loader = SheetImageLoader(sheet)
    img_frame_map = {}

    rows = sheet.max_row
    for row in range(2, rows + 1):
        name_cn = sheet.cell(row, 3).value
        frame_type = sheet.cell(row, 7).value
        img_frame_map[name_cn] = frame_type
    files = os.listdir(old_folder)
    for file in files:
        if os.path.exists(os.path.join(new_folder, "道具-"+file)):
            shutil.move(os.path.join(old_folder, file), os.path.join(backup_folder, file))
            continue
        frame_type = img_frame_map.get(file[:-4], None)
        if frame_type is None:
            continue
        image = Image.open(os.path.join(old_folder, file)).convert("RGBA")
        image_np = np.array(image)
        print(file, len(image_np), len(image_np[0]))
        if len(image_np) != 152 or len(image_np[0]) != 152:
            continue
        out_img = Image.open(frame_type + "框.png").convert("RGBA")
        # out_img = np.array(frame_img) # 160*160*4
        out_img.paste(image, (4,4), image)
        # out_img = np.uint8(out_img)
        # im = Image.fromarray(out_img)
        out_img.save(os.path.join(new_folder, "道具-"+file))

def replaceAllImage(old_folder, new_folder):
    if not os.path.exists(new_folder):
        os.mkdir(new_folder)
    files = os.listdir(old_folder)
    for file in files:
        if os.path.isdir(os.path.join(old_folder, file)):
            replaceAllImage(os.path.join(old_folder, file), os.path.join(new_folder, file))
            continue
        image = Image.open(os.path.join(old_folder, file)).convert("RGBA")
        image_np = np.array(image)
        if len(image_np) > 160 or len(image_np[0]) > 160:
            print(file, len(image_np), len(image_np[0]))
            continue
        for frame_type in ["铜", "银", "金", "彩"]:
            out_img = Image.open(frame_type + "框.png").convert("RGBA")
            left = (160 - len(image_np[0])) // 2
            top = (160 - len(image_np)) // 2
            out_img.paste(image, (left, top), image)
            out_img.save(os.path.join(new_folder, frame_type+file))

if __name__ == "__main__":
    # main()
    # getXlsxFile("徽章", "badge")
    # replaceImageName("道具图鉴无底")
    # replaceImage("道具图鉴", "道具图鉴无底", "道具图鉴有底", "道具图鉴无底\\备份")
    replaceAllImage("道具图鉴无底", "道具图鉴有底")